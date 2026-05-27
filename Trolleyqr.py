import streamlit as st
import pandas as pd
import qrcode
import io
import os
import json
import zipfile
import base64
from PIL import Image, ImageDraw, ImageFont
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ─── Page Config ────────────────────────────────────────────────────────────────
st.set_page_config(
    page_title="Trolley QR Manager",
    page_icon="🏭",
    layout="wide",
    initial_sidebar_state="expanded",
)

# ─── Detect QR Scan (mobile view) ───────────────────────────────────────────────
query_params = st.query_params
scanned_trolley = query_params.get("trolley", None)

# ─── CSS Styling ────────────────────────────────────────────────────────────────
MOBILE_CSS = """
<style>
@import url('https://fonts.googleapis.com/css2?family=DM+Sans:wght@400;500;600;700&family=DM+Mono:wght@400;500&display=swap');
* { box-sizing: border-box; margin: 0; padding: 0; }
body { background: #F5F5F0; font-family: 'DM Sans', sans-serif; }
.phone-wrap { max-width: 480px; margin: 0 auto; padding: 0; background: #fff; min-height: 100vh; }
.phone-header {
    background: #111; padding: 18px 20px 16px;
    position: sticky; top: 0; z-index: 10;
}
.phone-header h1 { color: #fff; font-size: 1.1rem; font-weight: 700; letter-spacing: 1px; }
.phone-header .sub { color: #888; font-size: 0.75rem; font-family: 'DM Mono', monospace; margin-top: 2px; }
.stats-strip { display: flex; background: #FAFAF8; border-bottom: 1px solid #E8E8E0; }
.stat { flex: 1; padding: 12px 8px; text-align: center; border-right: 1px solid #E8E8E0; }
.stat:last-child { border-right: none; }
.stat-n { font-size: 1.4rem; font-weight: 700; color: #111; font-family: 'DM Mono', monospace; }
.stat-l { font-size: 0.6rem; color: #999; letter-spacing: 1px; text-transform: uppercase; margin-top: 2px; }
.part-card {
    border-bottom: 1px solid #F0F0EA; padding: 14px 16px;
    background: #fff;
}
.part-card:nth-child(even) { background: #FAFAF8; }
.part-top { display: flex; justify-content: space-between; align-items: flex-start; margin-bottom: 8px; }
.part-no { font-family: 'DM Mono', monospace; font-size: 0.75rem; color: #111; font-weight: 500;
           background: #F0F0EA; padding: 2px 8px; border-radius: 4px; }
.part-desc { font-size: 0.9rem; color: #333; font-weight: 500; margin-top: 4px; }
.qty-block { text-align: right; }
.qty-big { font-size: 1.3rem; font-weight: 700; color: #111; font-family: 'DM Mono', monospace; }
.qty-label { font-size: 0.6rem; color: #999; letter-spacing: 1px; text-transform: uppercase; }
.pending-ok { font-size: 0.7rem; color: #16A34A; font-weight: 500; }
.pending-warn { font-size: 0.7rem; color: #D97706; font-weight: 500; }
.tags { display: flex; flex-wrap: wrap; gap: 4px; margin-top: 8px; }
.tag { background: #F0F0EA; border-radius: 4px; padding: 2px 7px;
       font-size: 0.68rem; color: #666; font-family: 'DM Mono', monospace; }
.section-label { padding: 8px 16px; background: #F5F5F0; font-size: 0.65rem;
                 color: #999; letter-spacing: 2px; text-transform: uppercase;
                 border-bottom: 1px solid #E8E8E0; font-family: 'DM Mono', monospace; }
.footer { padding: 24px 16px; text-align: center; color: #bbb; font-size: 0.75rem; }
</style>
"""

MAIN_CSS = """
<style>
@import url('https://fonts.googleapis.com/css2?family=DM+Sans:wght@400;500;600;700&family=DM+Mono:wght@400;500&display=swap');

:root {
    --accent: #111111;
    --accent2: #444;
    --bg: #F7F7F4;
    --card: #FFFFFF;
    --border: #E2E2DC;
    --muted: #888;
    --text: #111;
    --green: #16A34A;
    --amber: #D97706;
    --surface2: #F0F0EA;
}

html, body, [class*="css"] {
    font-family: 'DM Sans', sans-serif !important;
    background-color: var(--bg) !important;
    color: var(--text) !important;
}

.stApp { background: var(--bg) !important; }

.main-header {
    background: #111;
    border-radius: 8px;
    padding: 24px 32px;
    margin-bottom: 28px;
    display: flex; align-items: center; gap: 16px;
}
.main-header h1 {
    font-size: 1.8rem; font-weight: 700; color: #fff;
    margin: 0; letter-spacing: 1px;
}
.main-header p { color: #888; margin: 0; font-size: 0.85rem; font-family: 'DM Mono', monospace; }

.trolley-card {
    background: var(--card); border: 1px solid var(--border);
    border-left: 3px solid #111; border-radius: 6px;
    padding: 14px 18px; margin-bottom: 10px;
    display: flex; justify-content: space-between; align-items: center;
}
.trolley-name { font-size: 1rem; font-weight: 700; color: #111; letter-spacing: 1px; }
.trolley-meta { font-family: 'DM Mono', monospace; font-size: 0.72rem; color: var(--muted); margin-top: 2px; }
.badge {
    background: #F0F0EA; border: 1px solid #D0D0C8; color: #444;
    padding: 3px 10px; border-radius: 4px; font-size: 0.72rem;
    font-family: 'DM Mono', monospace; font-weight: 600;
}
.badge-green { background: #DCFCE7; border-color: #86EFAC; color: #16A34A; }
.stat-box {
    background: var(--card); border: 1px solid var(--border);
    border-top: 3px solid #111; border-radius: 6px;
    padding: 16px; text-align: center;
}
.stat-number { font-size: 2rem; font-weight: 700; color: #111; font-family: 'DM Mono', monospace; }
.stat-label { color: var(--muted); font-size: 0.75rem; letter-spacing: 1px; text-transform: uppercase; margin-top: 4px; }

section[data-testid="stSidebar"] { background: #FAFAF8 !important; border-right: 1px solid var(--border) !important; }

.stButton > button {
    background: #111 !important; color: #fff !important;
    font-family: 'DM Sans', sans-serif !important; font-weight: 600 !important;
    letter-spacing: 0.5px !important; border: none !important;
    border-radius: 6px !important; padding: 10px 24px !important;
}
.stButton > button:hover { background: #333 !important; }
.stDownloadButton > button {
    background: transparent !important; color: #16A34A !important;
    border: 1px solid #86EFAC !important; font-family: 'DM Sans', sans-serif !important;
    font-weight: 600 !important; border-radius: 6px !important;
}
.stTextInput > div > div > input {
    background: #fff !important; border: 1px solid var(--border) !important;
    color: var(--text) !important; border-radius: 6px !important;
    font-family: 'DM Mono', monospace !important;
}
.stTabs [data-baseweb="tab-list"] { background: transparent; border-bottom: 2px solid var(--border); gap: 0; }
.stTabs [data-baseweb="tab"] {
    font-family: 'DM Sans', sans-serif; font-weight: 600; letter-spacing: 0.5px;
    color: var(--muted) !important; background: transparent !important;
    border-bottom: 2px solid transparent !important; padding: 12px 24px;
}
.stTabs [aria-selected="true"] { color: #111 !important; border-bottom: 2px solid #111 !important; }

.section-title {
    font-size: 1rem; font-weight: 700; color: #111; letter-spacing: 0.5px;
    text-transform: uppercase; border-bottom: 1px solid var(--border);
    padding-bottom: 8px; margin: 20px 0 14px;
}
.step-label { font-family: 'DM Mono', monospace; font-size: 0.72rem; color: #888;
              letter-spacing: 1px; text-transform: uppercase; margin-bottom: 6px; }

.scan-instruction { background: var(--surface2); border: 1px solid var(--border);
    border-radius: 6px; padding: 16px; text-align: center; margin-top: 12px; }
.scan-instruction p { color: #555; font-size: 0.85rem; margin: 0; line-height: 1.5; }

.url-box { background: #F0F0EA; border: 1px solid #D8D8D0; border-radius: 6px;
           padding: 10px 14px; font-family: 'DM Mono', monospace; font-size: 0.72rem;
           color: #444; word-break: break-all; margin-top: 8px; }

hr { border-color: var(--border) !important; }
.stSuccess { background: #DCFCE7 !important; border: 1px solid #86EFAC !important; }
.stInfo { background: #F0F0EA !important; border: 1px solid #D0D0C8 !important; }
.stWarning { background: #FEF9C3 !important; border: 1px solid #FDE047 !important; }
</style>
"""

# ─── Data Storage ────────────────────────────────────────────────────────────────
if "trolleys" not in st.session_state:
    st.session_state.trolleys = {}

EXPECTED_COLS = [
    "S.No", "Part No", "Description", "Store", "Zone",
    "Rack", "Picking Location (old Location)",
    "Trolley Location", "Qty", "Pick Qty", "Pending Qty",
    "Delivery Location", "Family"
]

# ─── Helper Functions ────────────────────────────────────────────────────────────

def get_app_url() -> str:
    """Get the base URL of this Streamlit app."""
    try:
        # Works on Streamlit Cloud
        return st.get_option("browser.serverAddress") or "http://localhost:8501"
    except Exception:
        return "http://localhost:8501"


def build_qr_url(trolley_name: str) -> str:
    """
    Build the URL that encodes the trolley pick list.
    When deployed on Streamlit Cloud, this becomes the real hosted URL.
    Workers scan this and the pick list opens directly in their browser.
    """
    # Try to get the actual deployed URL from headers
    try:
        headers = st.context.headers
        host = headers.get("host", "localhost:8501")
        # Detect if running on HTTPS (Streamlit Cloud)
        proto = "https" if "streamlit.app" in host or "streamlitapp.com" in host else "http"
        base_url = f"{proto}://{host}"
    except Exception:
        base_url = "http://localhost:8501"
    
    return f"{base_url}/?trolley={trolley_name}"


def generate_qr_code(url: str, trolley_name: str) -> bytes:
    """Generate a black & white QR code with label."""
    qr = qrcode.QRCode(
        version=None,
        error_correction=qrcode.constants.ERROR_CORRECT_H,
        box_size=10,
        border=4,
    )
    qr.add_data(url)
    qr.make(fit=True)
    # Classic black on white
    qr_img = qr.make_image(fill_color="black", back_color="white").convert("RGB")

    label_height = 80
    total_height = qr_img.height + label_height
    canvas = Image.new("RGB", (qr_img.width, total_height), "white")
    canvas.paste(qr_img, (0, 0))

    draw = ImageDraw.Draw(canvas)
    try:
        font_big = ImageFont.truetype("/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf", 20)
        font_small = ImageFont.truetype("/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf", 13)
    except Exception:
        font_big = ImageFont.load_default()
        font_small = font_big

    # Draw black top border line
    draw.rectangle([(0, qr_img.height), (qr_img.width, qr_img.height + 2)], fill="#111")

    text = f"TROLLEY: {trolley_name.upper()}"
    bbox = draw.textbbox((0, 0), text, font=font_big)
    tw = bbox[2] - bbox[0]
    draw.text(((qr_img.width - tw) // 2, qr_img.height + 10), text, fill="#111111", font=font_big)

    sub = "Scan to view Pick List"
    bbox2 = draw.textbbox((0, 0), sub, font=font_small)
    sw = bbox2[2] - bbox2[0]
    draw.text(((qr_img.width - sw) // 2, qr_img.height + 42), sub, fill="#666666", font=font_small)

    buf = io.BytesIO()
    canvas.save(buf, format="PNG", dpi=(300, 300))
    return buf.getvalue()


def df_to_styled_excel(df: pd.DataFrame, trolley_name: str) -> bytes:
    buf = io.BytesIO()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = trolley_name[:31]

    ws.merge_cells(f"A1:{get_column_letter(len(df.columns))}1")
    title_cell = ws["A1"]
    title_cell.value = f"PICK LIST — TROLLEY: {trolley_name.upper()}"
    title_cell.font = Font(bold=True, size=14, color="FFFFFF")
    title_cell.fill = PatternFill("solid", fgColor="111111")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    header_fill = PatternFill("solid", fgColor="F0F0EA")
    header_font = Font(bold=True, color="111111", size=10)
    thin = Side(style="thin", color="D0D0C8")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col_idx, col_name in enumerate(df.columns, start=1):
        cell = ws.cell(row=2, column=col_idx, value=col_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border
    ws.row_dimensions[2].height = 36

    alt_fill = PatternFill("solid", fgColor="F7F7F4")
    white_fill = PatternFill("solid", fgColor="FFFFFF")
    data_font = Font(color="111111", size=9)
    center_align = Alignment(horizontal="center", vertical="center")

    for row_idx, row in enumerate(df.itertuples(index=False), start=3):
        fill = alt_fill if row_idx % 2 == 0 else white_fill
        for col_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = data_font
            cell.fill = fill
            cell.alignment = center_align
            cell.border = border
        ws.row_dimensions[row_idx].height = 20

    col_widths = {"S.No": 6, "Part No": 14, "Description": 28, "Store": 10,
                  "Zone": 8, "Rack": 8, "Picking Location (old Location)": 24,
                  "Trolley Location": 18, "Qty": 8, "Pick Qty": 10,
                  "Pending Qty": 12, "Delivery Location": 20, "Family": 14}
    for col_idx, col_name in enumerate(df.columns, start=1):
        width = col_widths.get(col_name, 14)
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.freeze_panes = "A3"
    wb.save(buf)
    return buf.getvalue()


def make_zip_all() -> bytes:
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w") as zf:
        for name in st.session_state.trolleys:
            url = build_qr_url(name)
            qr_bytes = generate_qr_code(url, name)
            zf.writestr(f"QR_{name}.png", qr_bytes)
    return buf.getvalue()


def read_uploaded_excel(uploaded_file) -> pd.DataFrame:
    df = pd.read_excel(uploaded_file, engine="openpyxl")
    df.columns = [str(c).strip() for c in df.columns]
    return df


# ════════════════════════════════════════════════════════════════════════════════
# MOBILE SCAN VIEW — shown when ?trolley=NAME is in URL
# ════════════════════════════════════════════════════════════════════════════════
if scanned_trolley and scanned_trolley in st.session_state.trolleys:
    st.markdown(MOBILE_CSS, unsafe_allow_html=True)

    df = st.session_state.trolleys[scanned_trolley]["df"]
    total_qty = int(df["Qty"].sum()) if "Qty" in df.columns and pd.api.types.is_numeric_dtype(df["Qty"]) else "–"
    pending_qty = int(df["Pending Qty"].sum()) if "Pending Qty" in df.columns and pd.api.types.is_numeric_dtype(df["Pending Qty"]) else "–"
    pick_qty = int(df["Pick Qty"].sum()) if "Pick Qty" in df.columns and pd.api.types.is_numeric_dtype(df["Pick Qty"]) else "–"

    st.markdown(f"""
    <div class="phone-wrap">
        <div class="phone-header">
            <h1>🏭 PICK LIST</h1>
            <div class="sub">TROLLEY: {scanned_trolley}</div>
        </div>
        <div class="stats-strip">
            <div class="stat"><div class="stat-n">{len(df)}</div><div class="stat-l">Parts</div></div>
            <div class="stat"><div class="stat-n">{total_qty}</div><div class="stat-l">Total Qty</div></div>
            <div class="stat"><div class="stat-n">{pick_qty}</div><div class="stat-l">Pick Qty</div></div>
            <div class="stat"><div class="stat-n">{pending_qty}</div><div class="stat-l">Pending</div></div>
        </div>
        <div class="section-label">Parts List · {len(df)} items</div>
    """, unsafe_allow_html=True)

    for i, row in df.iterrows():
        part_no = row.get("Part No", "—")
        desc = row.get("Description", "—")
        qty = row.get("Qty", "—")
        pick = row.get("Pick Qty", "—")
        pending = row.get("Pending Qty", "—")
        location = row.get("Picking Location (old Location)", "—")
        trolley_loc = row.get("Trolley Location", "—")
        zone = row.get("Zone", "—")
        rack = row.get("Rack", "—")
        family = row.get("Family", "—")
        delivery = row.get("Delivery Location", "—")

        try:
            pending_int = int(pending)
            pending_class = "pending-ok" if pending_int == 0 else "pending-warn"
            pending_icon = "✓ All picked" if pending_int == 0 else f"⚠ {pending_int} pending"
        except (ValueError, TypeError):
            pending_class = "pending-warn"
            pending_icon = f"{pending} pending"

        st.markdown(f"""
        <div class="part-card">
            <div class="part-top">
                <div>
                    <span class="part-no">{part_no}</span>
                    <div class="part-desc">{desc}</div>
                </div>
                <div class="qty-block">
                    <div class="qty-big">{qty}</div>
                    <div class="qty-label">QTY</div>
                    <div class="{pending_class}">{pending_icon}</div>
                </div>
            </div>
            <div class="tags">
                <span class="tag">📍 {location}</span>
                <span class="tag">🛒 {trolley_loc}</span>
                <span class="tag">Zone: {zone}</span>
                <span class="tag">Rack: {rack}</span>
                <span class="tag">Pick: {pick}</span>
                {'<span class="tag">📦 ' + str(family) + '</span>' if str(family) not in ['—','nan',''] else ''}
                {'<span class="tag">🚚 ' + str(delivery) + '</span>' if str(delivery) not in ['—','nan',''] else ''}
            </div>
        </div>
        """, unsafe_allow_html=True)

    st.markdown('<div class="footer">🏭 Trolley QR Manager · Scan complete</div></div>', unsafe_allow_html=True)

    excel_bytes = df_to_styled_excel(df, scanned_trolley)
    st.download_button(
        label="⬇️ Download Pick List as Excel",
        data=excel_bytes,
        file_name=f"PickList_{scanned_trolley}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    st.stop()

elif scanned_trolley:
    # Trolley not found in session (happens when deployed — need persistent storage)
    st.warning(f"⚠️ Trolley **{scanned_trolley}** not found in current session.")
    st.info("💡 **Deployment Note:** For QR scanning to work across sessions, connect a database (Google Sheets, Firebase, Supabase, etc.) so data persists. In local mode, the admin must keep the app open.")
    st.stop()


# ════════════════════════════════════════════════════════════════════════════════
# MAIN ADMIN UI
# ════════════════════════════════════════════════════════════════════════════════
st.markdown(MAIN_CSS, unsafe_allow_html=True)

# Header
st.markdown("""
<div class="main-header">
    <div>
        <h1>🏭 Trolley QR Manager</h1>
        <p>Manufacturing Pick-List System · Upload → Register → Generate QR → Scan on Phone</p>
    </div>
</div>
""", unsafe_allow_html=True)

# Stats
total_trolleys = len(st.session_state.trolleys)
total_parts = sum(len(v["df"]) for v in st.session_state.trolleys.values()) if total_trolleys else 0

c1, c2, c3 = st.columns(3)
with c1:
    st.markdown(f"""<div class="stat-box">
        <div class="stat-number">{total_trolleys}</div>
        <div class="stat-label">Trolleys Registered</div>
    </div>""", unsafe_allow_html=True)
with c2:
    st.markdown(f"""<div class="stat-box">
        <div class="stat-number">{total_parts}</div>
        <div class="stat-label">Total Parts</div>
    </div>""", unsafe_allow_html=True)
with c3:
    st.markdown(f"""<div class="stat-box">
        <div class="stat-number">{total_trolleys}</div>
        <div class="stat-label">QR Codes Ready</div>
    </div>""", unsafe_allow_html=True)

st.markdown("<br>", unsafe_allow_html=True)

tab1, tab2, tab3 = st.tabs(["📤  UPLOAD & GENERATE", "📋  VIEW PICK LISTS", "📱  SCAN SIMULATOR"])

# ── TAB 1 ─────────────────────────────────────────────────────────────────────
with tab1:
    col_left, col_right = st.columns([1, 1], gap="large")

    with col_left:
        st.markdown('<div class="section-title">Step 1 · Upload Excel & Name Trolley</div>', unsafe_allow_html=True)

        trolley_name = st.text_input(
            "Trolley Name / ID",
            placeholder="e.g. TROLLEY-A, LINE-3, BIN-07",
        ).strip().upper()

        uploaded_file = st.file_uploader("Upload Pick List Excel (.xlsx)", type=["xlsx", "xls"])

        st.markdown('<div class="step-label">Expected Columns (auto-mapped if present)</div>', unsafe_allow_html=True)
        st.caption(" · ".join(EXPECTED_COLS))

        if st.button("✅ Register Trolley & Generate QR", use_container_width=True):
            if not trolley_name:
                st.error("⚠️ Please enter a Trolley Name.")
            elif not uploaded_file:
                st.error("⚠️ Please upload an Excel file.")
            else:
                df = read_uploaded_excel(uploaded_file)
                for col in EXPECTED_COLS:
                    if col not in df.columns:
                        df[col] = ""
                ordered = [c for c in EXPECTED_COLS if c in df.columns]
                extras = [c for c in df.columns if c not in EXPECTED_COLS]
                df = df[ordered + extras]

                st.session_state.trolleys[trolley_name] = {
                    "df": df,
                    "filename": uploaded_file.name
                }
                st.success(f"✅ Trolley **{trolley_name}** registered with {len(df)} parts!")
                st.rerun()

        if st.session_state.trolleys:
            st.markdown('<div class="section-title">Registered Trolleys</div>', unsafe_allow_html=True)

            for name, data in st.session_state.trolleys.items():
                row_c1, row_c2 = st.columns([3, 1])
                with row_c1:
                    st.markdown(f"""
                    <div class="trolley-card">
                        <div>
                            <div class="trolley-name">{name}</div>
                            <div class="trolley-meta">{data['filename']} · {len(data['df'])} parts</div>
                        </div>
                        <span class="badge badge-green">READY</span>
                    </div>
                    """, unsafe_allow_html=True)
                with row_c2:
                    if st.button("🗑️", key=f"del_{name}", help=f"Remove {name}"):
                        del st.session_state.trolleys[name]
                        st.rerun()

            st.markdown("---")
            st.markdown('<div class="step-label">Bulk Download</div>', unsafe_allow_html=True)
            zip_bytes = make_zip_all()
            st.download_button(
                label="⬇️ Download ALL QR Codes (.zip)",
                data=zip_bytes,
                file_name="all_trolley_qr_codes.zip",
                mime="application/zip",
                use_container_width=True
            )

    with col_right:
        st.markdown('<div class="section-title">Step 2 · Preview & Download QR Code</div>', unsafe_allow_html=True)

        if st.session_state.trolleys:
            selected_qr = st.selectbox(
                "Select Trolley to Preview",
                options=list(st.session_state.trolleys.keys()),
                key="qr_preview_select"
            )

            if selected_qr:
                url = build_qr_url(selected_qr)
                qr_bytes = generate_qr_code(url, selected_qr)

                qr_img = Image.open(io.BytesIO(qr_bytes))
                st.image(qr_img, caption=f"QR Code for {selected_qr}", use_container_width=False, width=280)

                st.markdown('<div class="step-label">QR encodes this URL</div>', unsafe_allow_html=True)
                st.markdown(f'<div class="url-box">{url}</div>', unsafe_allow_html=True)

                st.markdown(f"""
                <div class="scan-instruction">
                    <p>📱 Print and paste this QR on <strong>{selected_qr}</strong>.<br>
                    Workers scan with any phone camera → pick list opens in browser instantly, no app needed.</p>
                </div>
                """, unsafe_allow_html=True)

                st.markdown("<br>", unsafe_allow_html=True)
                st.download_button(
                    label=f"⬇️ Download QR — {selected_qr}.png",
                    data=qr_bytes,
                    file_name=f"QR_{selected_qr}.png",
                    mime="image/png",
                    use_container_width=True
                )

                excel_bytes = df_to_styled_excel(
                    st.session_state.trolleys[selected_qr]["df"], selected_qr
                )
                st.download_button(
                    label=f"⬇️ Download Pick List Excel — {selected_qr}",
                    data=excel_bytes,
                    file_name=f"PickList_{selected_qr}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True
                )
        else:
            st.info("ℹ️ Register at least one trolley on the left to generate a QR code.")

# ── TAB 2 ─────────────────────────────────────────────────────────────────────
with tab2:
    st.markdown('<div class="section-title">All Pick Lists</div>', unsafe_allow_html=True)

    if not st.session_state.trolleys:
        st.info("ℹ️ No trolleys registered yet.")
    else:
        for name, data in st.session_state.trolleys.items():
            with st.expander(f"🛒  {name}  ·  {len(data['df'])} parts  ·  {data['filename']}", expanded=False):
                df = data["df"]

                m1, m2, m3 = st.columns(3)
                if "Qty" in df.columns:
                    m1.metric("Total Qty", int(df["Qty"].sum()) if pd.api.types.is_numeric_dtype(df["Qty"]) else "—")
                if "Pick Qty" in df.columns:
                    m2.metric("Pick Qty", int(df["Pick Qty"].sum()) if pd.api.types.is_numeric_dtype(df["Pick Qty"]) else "—")
                if "Pending Qty" in df.columns:
                    m3.metric("Pending", int(df["Pending Qty"].sum()) if pd.api.types.is_numeric_dtype(df["Pending Qty"]) else "—")

                st.dataframe(df, use_container_width=True, hide_index=True)

                excel_bytes = df_to_styled_excel(df, name)
                st.download_button(
                    label=f"⬇️ Download Excel — {name}",
                    data=excel_bytes,
                    file_name=f"PickList_{name}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    key=f"dl_excel_{name}"
                )

# ── TAB 3 ─────────────────────────────────────────────────────────────────────
with tab3:
    st.markdown('<div class="section-title">📱 Scan Simulator — What Workers See on Phone</div>', unsafe_allow_html=True)
    st.caption("Click the link below to simulate opening the scanned URL. This is exactly what a worker's phone browser shows.")

    if not st.session_state.trolleys:
        st.info("ℹ️ Register trolleys first.")
    else:
        sim_trolley = st.selectbox(
            "Simulate scanning QR for trolley:",
            options=list(st.session_state.trolleys.keys()),
            key="sim_select"
        )

        if sim_trolley:
            url = build_qr_url(sim_trolley)
            st.markdown(f"""
            <div class="url-box">🔗 {url}</div>
            """, unsafe_allow_html=True)
            st.markdown("<br>", unsafe_allow_html=True)

            # Open the scan view in same tab
            st.link_button(f"📱 Open Scan View for {sim_trolley}", url, use_container_width=True)

            st.markdown("---")
            st.info("👆 Click above to see exactly what workers see when they scan the QR code. The pick list opens as a clean mobile-friendly page.")

# ── Sidebar ────────────────────────────────────────────────────────────────────
with st.sidebar:
    st.markdown("""
    <div style="font-family:'DM Sans',sans-serif; font-size:1rem; font-weight:700;
                color:#111; letter-spacing:1px; text-transform:uppercase;
                border-bottom:1px solid #E2E2DC; padding-bottom:8px; margin-bottom:16px;">
        How It Works
    </div>
    """, unsafe_allow_html=True)

    steps = [
        ("1", "Upload Excel", "Prepare a .xlsx file with the pick list for each trolley."),
        ("2", "Name Trolley", "Enter the trolley ID (e.g. TROLLEY-A)."),
        ("3", "Generate QR", "Click Register. A black & white QR code is created."),
        ("4", "Download & Print", "Download the QR PNG and paste on the physical trolley."),
        ("5", "Worker Scans", "Worker scans with phone → pick list opens in browser, no app needed."),
    ]

    for num, title, desc in steps:
        st.markdown(f"""
        <div style="display:flex; gap:12px; margin-bottom:14px; align-items:flex-start;">
            <div style="background:#111; color:#fff; font-weight:700; font-size:0.75rem;
                        width:22px; height:22px; border-radius:50%; display:flex; align-items:center;
                        justify-content:center; flex-shrink:0; font-family:monospace;">{num}</div>
            <div>
                <div style="font-weight:700; font-size:0.85rem; color:#111;">{title}</div>
                <div style="font-size:0.75rem; color:#888; margin-top:2px; line-height:1.4;">{desc}</div>
            </div>
        </div>
        """, unsafe_allow_html=True)

    st.markdown("---")
    st.markdown("""
    <div style="font-size:0.75rem; color:#888; font-family:'DM Mono',monospace; line-height:1.7;
                background:#F0F0EA; border-radius:6px; padding:12px;">
        <strong style="color:#444;">⚠️ Persistent Storage</strong><br>
        On Streamlit Cloud, data resets between sessions.<br>
        For production, connect a database so QR scan data persists even after app restarts.
    </div>
    """, unsafe_allow_html=True)

    st.markdown("---")
    if st.session_state.trolleys and st.button("🗑️ Clear All Trolleys", use_container_width=True):
        st.session_state.trolleys = {}
        st.rerun()
